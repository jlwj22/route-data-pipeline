"""
Template manager for route pipeline reports.
Handles loading, parsing, and applying report templates.
"""

import json
import os
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.formatting import Rule

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.logger import get_logger

logger = get_logger(__name__)

class TemplateManager:
    """
    Manage and apply report templates to generate formatted reports.
    """
    
    def __init__(self, template_directory: str = None):
        if template_directory is None:
            template_directory = os.path.join(os.path.dirname(__file__), 'templates')
        
        self.template_directory = template_directory
        self.templates = {}
        self.load_templates()
        
    def load_templates(self):
        """Load all templates from the template directory."""
        logger.info(f"Loading templates from: {self.template_directory}")
        
        if not os.path.exists(self.template_directory):
            logger.warning(f"Template directory does not exist: {self.template_directory}")
            return
        
        template_files = [f for f in os.listdir(self.template_directory) 
                         if f.endswith('.json')]
        
        for template_file in template_files:
            template_path = os.path.join(self.template_directory, template_file)
            
            try:
                with open(template_path, 'r') as f:
                    template_data = json.load(f)
                
                template_name = template_data.get('template_name', 
                                                 os.path.splitext(template_file)[0])
                self.templates[template_name] = template_data
                
                logger.info(f"Loaded template: {template_name}")
                
            except Exception as e:
                logger.error(f"Error loading template {template_file}: {e}")
        
        logger.info(f"Loaded {len(self.templates)} templates")
    
    def get_template(self, template_name: str) -> Optional[Dict[str, Any]]:
        """Get a specific template by name."""
        return self.templates.get(template_name)
    
    def list_templates(self) -> List[str]:
        """List all available template names."""
        return list(self.templates.keys())
    
    def apply_template(self, template_name: str, data: Dict[str, Any], 
                      workbook: Workbook) -> Workbook:
        """
        Apply a template to a workbook with the provided data.
        """
        logger.info(f"Applying template: {template_name}")
        
        template = self.get_template(template_name)
        if not template:
            logger.error(f"Template not found: {template_name}")
            return workbook
        
        # Remove default sheet if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Process each sheet in the template
        for sheet_config in template.get('sheets', []):
            self._create_sheet_from_template(workbook, sheet_config, data, template)
        
        logger.info(f"Template {template_name} applied successfully")
        return workbook
    
    def _create_sheet_from_template(self, workbook: Workbook, 
                                   sheet_config: Dict[str, Any], 
                                   data: Dict[str, Any],
                                   template: Dict[str, Any]):
        """Create a worksheet from a sheet template configuration."""
        sheet_name = sheet_config.get('name', 'Sheet')
        sheet_type = sheet_config.get('type', 'data_table')
        
        logger.info(f"Creating sheet: {sheet_name} (type: {sheet_type})")
        
        # Create worksheet
        ws = workbook.create_sheet(sheet_name)
        
        # Apply sheet based on type
        if sheet_type == 'data_table':
            self._create_data_table_sheet(ws, sheet_config, data, template)
        elif sheet_type == 'kpi_dashboard':
            self._create_kpi_dashboard_sheet(ws, sheet_config, data, template)
        elif sheet_type == 'executive_summary':
            self._create_executive_summary_sheet(ws, sheet_config, data, template)
        elif sheet_type == 'charts':
            self._create_charts_sheet(ws, sheet_config, data, template)
        else:
            logger.warning(f"Unknown sheet type: {sheet_type}")
    
    def _create_data_table_sheet(self, worksheet, sheet_config: Dict[str, Any], 
                                data: Dict[str, Any], template: Dict[str, Any]):
        """Create a data table sheet."""
        columns = sheet_config.get('columns', [])
        sheet_data = data.get(sheet_config.get('data_source', 'routes'), [])
        
        # Create title
        title = f"{sheet_config.get('name')} - {data.get('report_date', 'Report')}"
        worksheet.merge_cells('A1:' + self._get_column_letter(len(columns)) + '1')
        worksheet['A1'] = title
        
        # Apply title styling
        self._apply_title_styling(worksheet['A1'], template.get('styling', {}))
        
        # Create headers
        for col_idx, column_config in enumerate(columns, 1):
            cell = worksheet.cell(row=3, column=col_idx, value=column_config.get('name'))
            self._apply_header_styling(cell, template.get('styling', {}))
            
            # Set column width
            column_letter = self._get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = column_config.get('width', 15)
        
        # Add data rows
        for row_idx, row_data in enumerate(sheet_data, 4):
            for col_idx, column_config in enumerate(columns, 1):
                field_name = column_config.get('field')
                cell_value = row_data.get(field_name, '')
                
                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Apply data formatting
                self._apply_data_formatting(cell, column_config, template.get('styling', {}))
        
        # Add summary rows
        summary_rows = sheet_config.get('summary_rows', [])
        if summary_rows and sheet_data:
            self._add_summary_rows(worksheet, summary_rows, columns, len(sheet_data), template)
        
        # Apply conditional formatting
        self._apply_conditional_formatting(worksheet, template.get('conditional_formatting', []), 
                                         columns, len(sheet_data))
    
    def _create_kpi_dashboard_sheet(self, worksheet, sheet_config: Dict[str, Any], 
                                   data: Dict[str, Any], template: Dict[str, Any]):
        """Create a KPI dashboard sheet."""
        metrics = sheet_config.get('metrics', [])
        
        # Create title
        title = f"{sheet_config.get('name')} - {data.get('report_date', 'Report')}"
        worksheet.merge_cells('A1:D1')
        worksheet['A1'] = title
        self._apply_title_styling(worksheet['A1'], template.get('styling', {}))
        
        # Add metrics
        for metric_config in metrics:
            position = metric_config.get('position', 'A3')
            metric_name = metric_config.get('name')
            field_name = metric_config.get('field')
            metric_type = metric_config.get('type', 'value')
            
            # Calculate metric value
            metric_value = self._calculate_metric_value(data, field_name, metric_type, metric_config)
            
            # Set metric name
            name_cell = worksheet[position]
            name_cell.value = metric_name
            self._apply_kpi_name_styling(name_cell, template.get('styling', {}))
            
            # Set metric value
            value_position = self._get_next_cell_position(position)
            value_cell = worksheet[value_position]
            value_cell.value = metric_value
            self._apply_kpi_value_styling(value_cell, template.get('styling', {}))
            
            # Apply number formatting
            if metric_config.get('format'):
                value_cell.number_format = metric_config['format']
    
    def _create_executive_summary_sheet(self, worksheet, sheet_config: Dict[str, Any], 
                                       data: Dict[str, Any], template: Dict[str, Any]):
        """Create an executive summary sheet."""
        sections = sheet_config.get('sections', [])
        
        # Create title
        title = f"Executive Summary - {data.get('report_date', 'Report')}"
        worksheet.merge_cells('A1:F1')
        worksheet['A1'] = title
        self._apply_title_styling(worksheet['A1'], template.get('styling', {}))
        
        # Add sections
        for section_config in sections:
            section_name = section_config.get('name')
            position = section_config.get('position', 'A3')
            metrics = section_config.get('metrics', [])
            
            # Add section header
            section_cell = worksheet[position]
            section_cell.value = section_name
            self._apply_section_header_styling(section_cell, template.get('styling', {}))
            
            # Add metrics for this section
            current_row = self._get_row_from_position(position) + 1
            for metric_config in metrics:
                metric_name = metric_config.get('name')
                field_name = metric_config.get('field')
                metric_type = metric_config.get('type', 'value')
                
                # Calculate metric value
                metric_value = self._calculate_metric_value(data, field_name, metric_type, metric_config)
                
                # Set metric name and value
                name_cell = worksheet.cell(row=current_row, column=1, value=metric_name)
                value_cell = worksheet.cell(row=current_row, column=2, value=metric_value)
                
                self._apply_executive_summary_styling(name_cell, value_cell, template.get('styling', {}))
                
                # Apply number formatting
                if metric_config.get('format'):
                    value_cell.number_format = metric_config['format']
                
                current_row += 1
    
    def _create_charts_sheet(self, worksheet, sheet_config: Dict[str, Any], 
                            data: Dict[str, Any], template: Dict[str, Any]):
        """Create a charts sheet."""
        charts = sheet_config.get('charts', [])
        
        # Create title
        title = f"Charts & Visualizations - {data.get('report_date', 'Report')}"
        worksheet.merge_cells('A1:P1')
        worksheet['A1'] = title
        self._apply_title_styling(worksheet['A1'], template.get('styling', {}))
        
        # Add placeholder text for charts
        for chart_config in charts:
            chart_name = chart_config.get('name')
            position = chart_config.get('position', 'A3')
            
            # Add chart placeholder
            chart_cell = worksheet[position]
            chart_cell.value = f"[{chart_name} Chart]"
            self._apply_chart_placeholder_styling(chart_cell, template.get('styling', {}))
            
            # Note: Actual chart creation would require chart_creator integration
            logger.info(f"Chart placeholder created: {chart_name} at {position}")
    
    def _apply_title_styling(self, cell, styling: Dict[str, Any]):
        """Apply title styling to a cell."""
        title_style = styling.get('title', {})
        
        if 'font' in title_style:
            font_config = title_style['font']
            cell.font = Font(
                size=font_config.get('size', 16),
                bold=font_config.get('bold', True),
                color=font_config.get('color', '000000')
            )
        
        if 'alignment' in title_style:
            align_config = title_style['alignment']
            cell.alignment = Alignment(
                horizontal=align_config.get('horizontal', 'center'),
                vertical=align_config.get('vertical', 'center')
            )
        
        if 'fill' in title_style:
            fill_config = title_style['fill']
            cell.fill = PatternFill(
                start_color=fill_config.get('color', 'FFFFFF'),
                end_color=fill_config.get('color', 'FFFFFF'),
                fill_type='solid'
            )
    
    def _apply_header_styling(self, cell, styling: Dict[str, Any]):
        """Apply header styling to a cell."""
        header_style = styling.get('header', {})
        
        if 'font' in header_style:
            font_config = header_style['font']
            cell.font = Font(
                bold=font_config.get('bold', True),
                color=font_config.get('color', 'FFFFFF')
            )
        
        if 'fill' in header_style:
            fill_config = header_style['fill']
            cell.fill = PatternFill(
                start_color=fill_config.get('color', '366092'),
                end_color=fill_config.get('color', '366092'),
                fill_type='solid'
            )
        
        if 'alignment' in header_style:
            align_config = header_style['alignment']
            cell.alignment = Alignment(
                horizontal=align_config.get('horizontal', 'center'),
                vertical=align_config.get('vertical', 'center')
            )
        
        # Apply border
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_data_formatting(self, cell, column_config: Dict[str, Any], 
                              styling: Dict[str, Any]):
        """Apply data formatting to a cell."""
        data_style = styling.get('data', {})
        
        # Apply number format
        if column_config.get('format'):
            cell.number_format = column_config['format']
        
        # Apply font
        if 'font' in data_style:
            font_config = data_style['font']
            cell.font = Font(size=font_config.get('size', 10))
        
        # Apply border
        if 'border' in data_style:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def _apply_kpi_name_styling(self, cell, styling: Dict[str, Any]):
        """Apply KPI name styling."""
        kpi_style = styling.get('kpi', {})
        
        if 'font' in kpi_style:
            font_config = kpi_style['font']
            cell.font = Font(
                size=font_config.get('size', 12),
                bold=font_config.get('bold', True)
            )
    
    def _apply_kpi_value_styling(self, cell, styling: Dict[str, Any]):
        """Apply KPI value styling."""
        kpi_style = styling.get('kpi', {})
        
        if 'font' in kpi_style:
            font_config = kpi_style['font']
            cell.font = Font(
                size=font_config.get('size', 12),
                bold=font_config.get('bold', True)
            )
        
        if 'fill' in kpi_style:
            fill_config = kpi_style['fill']
            cell.fill = PatternFill(
                start_color=fill_config.get('color', 'F2F2F2'),
                end_color=fill_config.get('color', 'F2F2F2'),
                fill_type='solid'
            )
    
    def _apply_section_header_styling(self, cell, styling: Dict[str, Any]):
        """Apply section header styling."""
        cell.font = Font(size=14, bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    def _apply_executive_summary_styling(self, name_cell, value_cell, styling: Dict[str, Any]):
        """Apply executive summary styling."""
        exec_style = styling.get('executive_summary', {})
        
        if 'font' in exec_style:
            font_config = exec_style['font']
            font = Font(size=font_config.get('size', 12))
            name_cell.font = font
            value_cell.font = font
        
        if 'fill' in exec_style:
            fill_config = exec_style['fill']
            fill = PatternFill(
                start_color=fill_config.get('color', 'F8F9FA'),
                end_color=fill_config.get('color', 'F8F9FA'),
                fill_type='solid'
            )
            name_cell.fill = fill
            value_cell.fill = fill
    
    def _apply_chart_placeholder_styling(self, cell, styling: Dict[str, Any]):
        """Apply chart placeholder styling."""
        cell.font = Font(size=12, bold=True, color='666666')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _add_summary_rows(self, worksheet, summary_rows: List[Dict], 
                         columns: List[Dict], data_row_count: int, template: Dict[str, Any]):
        """Add summary rows to the worksheet."""
        start_row = data_row_count + 5  # Leave some space after data
        
        for summary_idx, summary_config in enumerate(summary_rows):
            current_row = start_row + summary_idx
            
            # Add summary label
            label_cell = worksheet.cell(row=current_row, column=1, value=summary_config.get('label'))
            label_cell.font = Font(bold=True)
            
            # Add calculations
            calculations = summary_config.get('calculations', [])
            for calc_config in calculations:
                column_field = calc_config.get('column')
                function = calc_config.get('function', 'SUM')
                
                # Find the column index for this field
                col_idx = None
                for idx, col_config in enumerate(columns, 1):
                    if col_config.get('field') == column_field:
                        col_idx = idx
                        break
                
                if col_idx:
                    # Create formula
                    column_letter = self._get_column_letter(col_idx)
                    formula = f"={function}({column_letter}4:{column_letter}{data_row_count + 3})"
                    
                    calc_cell = worksheet.cell(row=current_row, column=col_idx, value=formula)
                    calc_cell.font = Font(bold=True)
                    
                    # Apply appropriate formatting
                    column_config = columns[col_idx - 1]
                    if column_config.get('format'):
                        calc_cell.number_format = column_config['format']
    
    def _apply_conditional_formatting(self, worksheet, conditional_formats: List[Dict], 
                                    columns: List[Dict], data_row_count: int):
        """Apply conditional formatting to the worksheet."""
        for format_config in conditional_formats:
            range_type = format_config.get('range')
            format_type = format_config.get('type')
            
            # Find the column for this formatting
            col_idx = None
            for idx, col_config in enumerate(columns, 1):
                if col_config.get('field') in range_type:
                    col_idx = idx
                    break
            
            if not col_idx:
                continue
            
            # Create the range
            column_letter = self._get_column_letter(col_idx)
            range_str = f"{column_letter}4:{column_letter}{data_row_count + 3}"
            
            # Apply formatting based on type
            if format_type == 'color_scale':
                colors = format_config.get('colors', ['FF0000', 'FFFF00', '00FF00'])
                rule = ColorScaleRule(
                    start_type='min', start_color=colors[0],
                    mid_type='percentile', mid_value=50, mid_color=colors[1],
                    end_type='max', end_color=colors[2]
                )
                worksheet.conditional_formatting.add(range_str, rule)
            
            elif format_type == 'data_bars':
                color = format_config.get('color', '4ECDC4')
                rule = DataBarRule(
                    start_type='min', start_value=0,
                    end_type='max', end_value=100,
                    color=color
                )
                worksheet.conditional_formatting.add(range_str, rule)
    
    def _calculate_metric_value(self, data: Dict[str, Any], field_name: str, 
                              metric_type: str, metric_config: Dict[str, Any]) -> Any:
        """Calculate a metric value based on the configuration."""
        if metric_type == 'calculated':
            # Handle calculated metrics with formulas
            formula = metric_config.get('formula', '')
            # This would need to be implemented based on the specific formula logic
            return 0  # Placeholder
        
        elif metric_type == 'count':
            # Count items
            data_source = data.get('routes', [])
            return len(data_source)
        
        elif metric_type == 'sum':
            # Sum values
            data_source = data.get('routes', [])
            return sum(item.get(field_name, 0) for item in data_source)
        
        elif metric_type == 'average':
            # Average values
            data_source = data.get('routes', [])
            values = [item.get(field_name, 0) for item in data_source]
            return sum(values) / len(values) if values else 0
        
        else:
            # Direct value
            return data.get(field_name, 0)
    
    def _get_column_letter(self, column_index: int) -> str:
        """Convert column index to Excel column letter."""
        result = ""
        while column_index > 0:
            column_index -= 1
            result = chr(column_index % 26 + ord('A')) + result
            column_index //= 26
        return result
    
    def _get_next_cell_position(self, position: str) -> str:
        """Get the next cell position (one column to the right)."""
        column = position[0]
        row = position[1:]
        next_column = chr(ord(column) + 1)
        return f"{next_column}{row}"
    
    def _get_row_from_position(self, position: str) -> int:
        """Extract row number from cell position."""
        return int(''.join(filter(str.isdigit, position)))
    
    def create_custom_template(self, template_name: str, template_config: Dict[str, Any]) -> bool:
        """Create a custom template and save it to the templates directory."""
        try:
            template_path = os.path.join(self.template_directory, f"{template_name}.json")
            
            with open(template_path, 'w') as f:
                json.dump(template_config, f, indent=2)
            
            # Add to loaded templates
            self.templates[template_name] = template_config
            
            logger.info(f"Custom template created: {template_name}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating custom template {template_name}: {e}")
            return False
    
    def validate_template(self, template_config: Dict[str, Any]) -> List[str]:
        """Validate a template configuration and return any errors."""
        errors = []
        
        if 'template_name' not in template_config:
            errors.append("Template name is required")
        
        if 'sheets' not in template_config:
            errors.append("Sheets configuration is required")
        
        for sheet_idx, sheet_config in enumerate(template_config.get('sheets', [])):
            if 'name' not in sheet_config:
                errors.append(f"Sheet {sheet_idx}: name is required")
            
            if 'type' not in sheet_config:
                errors.append(f"Sheet {sheet_idx}: type is required")
            
            # Additional validation based on sheet type
            sheet_type = sheet_config.get('type')
            if sheet_type == 'data_table' and 'columns' not in sheet_config:
                errors.append(f"Sheet {sheet_idx}: columns are required for data_table type")
        
        return errors